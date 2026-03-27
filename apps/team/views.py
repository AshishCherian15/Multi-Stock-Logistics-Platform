from django.shortcuts import render
from django.contrib.auth.models import User
from django.db import models
from django.core.paginator import Paginator
from permissions.decorators import get_user_role, require_role
from .models import AccessRequest

@require_role('superadmin', 'admin')
def team_list(request):
    # Get all staff and superusers
    team_users = User.objects.filter(models.Q(is_staff=True) | models.Q(is_superuser=True)).distinct()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        team_users = team_users.filter(
            models.Q(username__icontains=search) |
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(email__icontains=search)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        team_users = team_users.filter(is_active=True)
    elif status_filter == 'inactive':
        team_users = team_users.filter(is_active=False)
    
    # Filter by role
    role_filter = request.GET.get('role', '')
    if role_filter == 'superadmin':
        team_users = team_users.filter(is_superuser=True)
    elif role_filter in ['admin', 'subadmin', 'staff']:
        team_users = team_users.filter(models.Q(role__role=role_filter) | models.Q(is_staff=True))
    
    team_users = team_users.order_by('-date_joined')
    total_count = team_users.count()
    
    # Pagination
    paginator = Paginator(team_users, 25)
    page_num = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_num)
    
    for user in page_obj:
        user.user_role = get_user_role(user)
    
    # Get access requests data
    pending_requests = AccessRequest.objects.filter(status='pending').select_related('user').order_by('-requested_at')
    active_users = User.objects.filter(is_active=True, is_staff=True).order_by('-date_joined')
    inactive_users = User.objects.filter(is_active=False, is_staff=True).order_by('-date_joined')
    
    for user in active_users:
        user.user_role = get_user_role(user)
    for user in inactive_users:
        user.user_role = get_user_role(user)
    
    context = {
        'page_obj': page_obj,
        'total_count': total_count,
        'pending_requests': pending_requests,
        'pending_count': pending_requests.count(),
        'active_users': active_users,
        'inactive_users': inactive_users,
    }
    
    return render(request, 'team_management.html', context)

@require_role('superadmin', 'admin')
def toggle_user_status(request, user_id):
    from django.http import JsonResponse
    try:
        user = User.objects.get(id=user_id)
        user.is_active = not user.is_active
        user.save()
        return JsonResponse({'success': True, 'is_active': user.is_active})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'})

@require_role('superadmin', 'admin')
def approve_access_request(request, request_id):
    from django.http import JsonResponse
    from django.utils import timezone
    try:
        access_req = AccessRequest.objects.get(id=request_id)
        access_req.status = 'approved'
        access_req.reviewed_by = request.user
        access_req.reviewed_at = timezone.now()
        access_req.save()
        
        user = access_req.user
        user.is_staff = True
        user.save()
        
        return JsonResponse({'success': True, 'message': 'Access request approved'})
    except AccessRequest.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Request not found'})

@require_role('superadmin', 'admin')
def reject_access_request(request, request_id):
    from django.http import JsonResponse
    from django.utils import timezone
    try:
        access_req = AccessRequest.objects.get(id=request_id)
        access_req.status = 'rejected'
        access_req.reviewed_by = request.user
        access_req.reviewed_at = timezone.now()
        access_req.save()
        
        return JsonResponse({'success': True, 'message': 'Access request rejected'})
    except AccessRequest.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Request not found'})

@require_role('superadmin', 'admin')
def create_user(request):
    from django.http import JsonResponse
    from permissions.models import UserRole
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        try:
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                password=data['password'],
                is_staff=True
            )
            UserRole.objects.create(user=user, role=data.get('role', 'staff'))
            return JsonResponse({'success': True, 'message': 'User created successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@require_role('superadmin', 'admin')
def update_user(request, user_id):
    from django.http import JsonResponse
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        try:
            user = User.objects.get(id=user_id)
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            user.email = data.get('email', user.email)
            user.save()
            return JsonResponse({'success': True, 'message': 'User updated successfully'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@require_role('superadmin', 'admin')
def change_role(request, user_id):
    from django.http import JsonResponse
    from permissions.models import UserRole
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        try:
            user = User.objects.get(id=user_id)
            role_obj, created = UserRole.objects.get_or_create(user=user)
            role_obj.role = data['role']
            role_obj.save()
            return JsonResponse({'success': True, 'message': 'Role updated successfully'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@require_role('superadmin', 'admin')
def delete_user(request, user_id):
    from django.http import JsonResponse
    if request.method == 'POST':
        try:
            user = User.objects.get(id=user_id)
            user.delete()
            return JsonResponse({'success': True, 'message': 'User deleted successfully'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'User not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@require_role('superadmin', 'admin')
def invite_user(request):
    from django.http import JsonResponse
    return JsonResponse({'success': True, 'message': 'Invite feature coming soon'})

@require_role('superadmin', 'admin')
def import_users(request):
    from django.http import JsonResponse
    return JsonResponse({'success': True, 'message': 'Import feature coming soon'})

@require_role('superadmin', 'admin')
def export_team(request):
    from django.http import HttpResponse
    import csv
    from datetime import datetime
    
    format_type = request.GET.get('format', 'csv')
    users = User.objects.filter(models.Q(is_staff=True) | models.Q(is_superuser=True)).distinct()
    
    if format_type == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Team Members'
            
            header_fill = PatternFill(start_color='0014A8', end_color='0014A8', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            headers = ['S.No', 'Username', 'First Name', 'Last Name', 'Email', 'Role', 'Status', 'Date Joined']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            for idx, user in enumerate(users, 1):
                role = get_user_role(user)
                ws.append([
                    idx,
                    user.username,
                    user.first_name or 'N/A',
                    user.last_name or 'N/A',
                    user.email,
                    role.title() if role else 'Staff',
                    'Active' if user.is_active else 'Inactive',
                    user.date_joined.strftime('%Y-%m-%d %H:%M')
                ])
            
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="team_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            format_type = 'csv'
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="team_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['S.No', 'Username', 'First Name', 'Last Name', 'Email', 'Role', 'Status', 'Date Joined'])
    
    for idx, user in enumerate(users, 1):
        role = get_user_role(user)
        writer.writerow([
            idx,
            user.username,
            user.first_name or 'N/A',
            user.last_name or 'N/A',
            user.email,
            role.title() if role else 'Staff',
            'Active' if user.is_active else 'Inactive',
            user.date_joined.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response
