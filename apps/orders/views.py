from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count
from django.views.decorators.http import require_http_methods
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from .models import Order, OrderItem
from goods.models import ListModel as Product
from customer.models import ListModel as Customer
from supplier.models import ListModel as Supplier
import uuid
import json
from datetime import datetime

# Import RBAC decorators
from permissions.decorators import require_permission, require_role, get_user_role

def get_filtered_orders(request, base_queryset):
    """
    Helper function to filter orders based on user role.
    SuperAdmin sees all orders, others see only their store's orders.
    """
    user = request.user
    if user.is_superuser:
        return base_queryset
    
    try:
        role = get_user_role(user)
        if role == 'superadmin':
            return base_queryset
    except:
        pass
    
    # Filter by user's openid
    openid = getattr(user, 'openid', user.username)
    return base_queryset.filter(openid=openid)

@login_required
def sales_orders(request):
    """Sales orders view - SuperAdmin/Admin full visibility; customers redirected."""
    from permissions.decorators import get_user_role
    from django.db.models import Prefetch
    role = get_user_role(request.user)
    
    # Redirect customers to their own orders page
    if role == 'customer':
        return redirect('orders:my_orders')
    
    try:
        # Base queryset with prefetch_related for items and customer
        base_qs = Order.objects.filter(order_type='sale').select_related('customer_user', 'customer').prefetch_related('items')
        orders = get_filtered_orders(request, base_qs).order_by('-created_at')
        
        # Apply filters
        status_filter = request.GET.get('status')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        
        if status_filter and status_filter != 'all':
            orders = orders.filter(status=status_filter)
        if from_date:
            orders = orders.filter(created_at__date__gte=from_date)
        if to_date:
            orders = orders.filter(created_at__date__lte=to_date)
        
        # Get base counts (also filtered by store)
        base_orders = get_filtered_orders(request, Order.objects.filter(order_type='sale'))
        
        context = {
            'orders': orders[:100],  # Limit to 100 for performance
            'total_orders': base_orders.count(),
            'pending_orders': base_orders.filter(status='pending').count(),
            'completed_orders': base_orders.filter(status='delivered').count(),
            'status_choices': Order.STATUS_CHOICES,
            'current_status': status_filter,
            'current_from_date': from_date,
            'current_to_date': to_date,
        }
        return render(request, 'admin/orders/list.html', context)
    except Exception as e:
        messages.error(request, f'Error loading orders: {str(e)}')
        return render(request, 'admin/orders/list.html', {'orders': [], 'total_orders': 0, 'pending_orders': 0, 'completed_orders': 0, 'status_choices': Order.STATUS_CHOICES})

@require_permission('orders', 'view')  # Orders view permission
def purchase_orders(request):
    """Purchase orders - customers should use my_orders instead"""
    from permissions.decorators import get_user_role
    role = get_user_role(request.user)
    
    # Redirect customers to their own orders page
    if role == 'customer':
        return redirect('orders:my_orders')
    orders = get_filtered_orders(request, Order.objects.filter(order_type='purchase')).order_by('-created_at')
    base_orders = get_filtered_orders(request, Order.objects.filter(order_type='purchase'))
    context = {
        'orders': orders,
        'total_orders': base_orders.count(),
        'pending_orders': base_orders.filter(status='pending').count(),
        'completed_orders': base_orders.filter(status='delivered').count(),
    }
    return render(request, 'orders/purchase_orders.html', context)

@require_permission('orders', 'create')  # Orders create permission
def create_order(request):
    """Create order - customers should use checkout instead"""
    from permissions.decorators import get_user_role
    role = get_user_role(request.user)
    
    # Redirect customers to checkout
    if role == 'customer':
        return redirect('cart:cart_page')
    
    if request.method == 'POST':
        order_type = request.POST.get('order_type')
        customer_id = request.POST.get('customer')
        supplier_id = request.POST.get('supplier')
        
        # Set openid for new order
        openid = getattr(request.user, 'openid', request.user.username)
        
        order = Order.objects.create(
            order_number=f"ORD-{uuid.uuid4().hex[:8].upper()}",
            order_type=order_type,
            customer_id=customer_id if customer_id else None,
            supplier_id=supplier_id if supplier_id else None,
            openid=openid,
            created_by=request.user
        )
        
        messages.success(request, f'Order {order.order_number} created successfully!')
        return redirect('orders:sales_orders' if order_type == 'sale' else 'orders:purchase_orders')
    
    # Filter products/customers/suppliers by store for non-SuperAdmin
    if request.user.is_superuser or get_user_role(request.user) == 'superadmin':
        customers = Customer.objects.all()
        suppliers = Supplier.objects.all()
        products = Product.objects.all()
    else:
        openid = getattr(request.user, 'openid', request.user.username)
        customers = Customer.objects.filter(openid=openid)
        suppliers = Supplier.objects.filter(openid=openid)
        products = Product.objects.filter(openid=openid)
    
    context = {
        'customers': customers,
        'suppliers': suppliers,
        'products': products,
    }
    return render(request, 'orders/create_order.html', context)

@login_required
def order_detail(request, order_id):
    from permissions.decorators import get_user_role
    
    order = get_object_or_404(Order, id=order_id)
    
    # Check permissions
    role = get_user_role(request.user)
    if role == 'customer':
        # Customers can only view their own orders
        if order.customer_user != request.user:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('You do not have permission to view this order')
    else:
        # Staff need orders view permission
        from permissions.decorators import check_permission
        if not check_permission(request.user, 'orders', 'view'):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('You do not have permission to view orders')
    
    items = OrderItem.objects.filter(order=order)
    context = {
        'order': order,
        'items': items,
    }
    return render(request, 'admin/orders/detail.html', context)

@require_permission('orders', 'approve')  # Order status change (approve/cancel)
def update_order_status(request, order_id):
    """Update order status - customers cannot update"""
    from permissions.decorators import get_user_role
    role = get_user_role(request.user)
    
    # Customers cannot update order status
    if role == 'customer':
        return JsonResponse({'success': False, 'error': 'Customers cannot update order status'}, status=403)
    if request.method == 'POST':
        from stock.models import StockListModel, StockMovement
        
        order = get_object_or_404(Order, id=order_id)
        old_status = order.status
        new_status = request.POST.get('status')
        
        # Restore stock if order is cancelled
        if new_status == 'cancelled' and old_status != 'cancelled' and order.order_type == 'sale':
            for item in order.items.all():
                if item.product:
                    stock = StockListModel.objects.filter(goods_code=item.product.goods_code).first()
                    if stock:
                        stock.goods_qty += item.quantity
                        stock.can_order_stock += item.quantity
                        stock.ordered_stock -= item.quantity
                        stock.save()
                        
                        StockMovement.objects.create(
                            goods_code=item.product.goods_code,
                            movement_type='in',
                            quantity=item.quantity,
                            reason=f'Order {order.order_number} cancelled',
                            user=request.user
                        )
            
            from notifications.email_service import send_order_cancelled
            if order.customer_user:
                send_order_cancelled(order)
        
        order.status = new_status
        order.save()
        
        # Log status change
        from .models import OrderStatusHistory
        OrderStatusHistory.objects.create(
            order=order,
            old_status=old_status,
            new_status=new_status,
            changed_by=request.user
        )
        
        messages.success(request, f'Order status updated to {order.get_status_display()}')
    return redirect('orders:order_detail', order_id=order_id)

@require_permission('orders', 'delete')  # Delete orders (SuperAdmin/Admin only)
def delete_order(request, order_id):
    """Delete order - Only SuperAdmin and Admin can delete"""
    from permissions.decorators import get_user_role
    role = get_user_role(request.user)
    
    # Only SuperAdmin and Admin can delete orders
    if role not in ['superadmin', 'admin']:
        messages.error(request, 'Permission denied. Only Admin and SuperAdmin can delete orders.')
        return redirect('orders:sales_orders')
    from stock.models import StockListModel, StockMovement
    
    order = get_object_or_404(Order, id=order_id)
    order_type = order.order_type
    order_number = order.order_number
    
    # Restore stock for sale orders
    if order_type == 'sale':
        for item in order.items.all():
            if item.product:
                stock = StockListModel.objects.filter(goods_code=item.product.goods_code).first()
                if stock:
                    stock.goods_qty += item.quantity
                    stock.can_order_stock += item.quantity
                    stock.ordered_stock -= item.quantity
                    stock.save()
                    
                    # Log stock movement
                    StockMovement.objects.create(
                        goods_code=item.product.goods_code,
                        movement_type='in',
                        quantity=item.quantity,
                        reason=f'Order {order_number} cancelled',
                        user=request.user
                    )
    
    order.delete()
    messages.success(request, f'Order {order_number} deleted successfully!')
    return redirect('orders:sales_orders' if order_type == 'sale' else 'orders:purchase_orders')

@require_permission('orders', 'create')  # Create order from cart (customers)
@require_http_methods(["POST"])
def create_order_from_cart(request):
    from cart.models import Cart
    from billing.models import Invoice, InvoiceItem
    from stock.models import StockListModel, StockMovement
    from datetime import timedelta
    from django.utils import timezone
    
    data = json.loads(request.body)
    cart = Cart.objects.filter(user=request.user).first()
    
    if not cart or cart.items.count() == 0:
        return JsonResponse({'success': False, 'message': 'Cart is empty'})
    
    # Check stock availability
    for cart_item in cart.items.all():
        if cart_item.product:
            stock = StockListModel.objects.filter(goods_code=cart_item.product.goods_code).first()
            if stock and stock.can_order_stock < cart_item.quantity:
                return JsonResponse({'success': False, 'message': f'Insufficient stock for {cart_item.product_name}'})
    
    # Create order
    order = Order.objects.create(
        order_type='sale',
        customer_user=request.user,
        status='confirmed',
        payment_status='unpaid',
        delivery_address=data.get('address', ''),
        delivery_phone=data.get('phone', ''),
        payment_method=data.get('payment_method', 'cod'),
        created_by=request.user
    )
    
    # Add items from cart and reduce stock
    total = 0
    for cart_item in cart.items.all():
        OrderItem.objects.create(
            order=order,
            product=cart_item.product,
            product_name=cart_item.product_name,
            quantity=cart_item.quantity,
            unit_price=cart_item.unit_price,
            seller=cart_item.seller
        )
        total += float(cart_item.total_price)
        
        # Reduce stock
        if cart_item.product:
            stock = StockListModel.objects.filter(goods_code=cart_item.product.goods_code).first()
            if stock:
                stock.goods_qty -= cart_item.quantity
                stock.can_order_stock -= cart_item.quantity
                stock.ordered_stock += cart_item.quantity
                stock.save()
                
                # Log stock movement
                StockMovement.objects.create(
                    goods_code=cart_item.product.goods_code,
                    movement_type='out',
                    quantity=cart_item.quantity,
                    reason=f'Order {order.order_number}',
                    user=request.user
                )
                
                # Check low stock
                from stock.utils import check_low_stock
                check_low_stock(cart_item.product.goods_code)
    
    order.total_amount = total
    order.discount_amount = float(data.get('discount', 0))
    order.save()
    
    # Auto-generate invoice
    invoice = Invoice.objects.create(
        invoice_type='product',
        customer_name=request.user.get_full_name() or request.user.username,
        customer_email=request.user.email,
        customer_phone=data.get('phone', ''),
        customer_address=data.get('address', ''),
        due_date=timezone.now().date() + timedelta(days=30),
        tax_rate=18,
        discount_amount=order.discount_amount,
        status='sent',
        created_by=request.user
    )
    
    # Add invoice items
    for item in order.items.all():
        InvoiceItem.objects.create(
            invoice=invoice,
            description=item.product_name,
            quantity=item.quantity,
            price=item.unit_price
        )
    
    invoice.subtotal = total
    invoice.tax_amount = (total * 18) / 100
    invoice.grand_total = total + invoice.tax_amount - float(invoice.discount_amount)
    invoice.save()
    
    # Clear cart
    cart.items.all().delete()
    
    # Send confirmation email
    from notifications.email_service import send_order_confirmation
    send_order_confirmation(order)
    
    return JsonResponse({
        'success': True,
        'order_number': order.order_number,
        'invoice_number': invoice.invoice_number,
        'order_id': order.id,
        'invoice_id': invoice.id,
        'redirect_to_payment': True
    })

@login_required
def my_orders(request):
    """Customers view their own orders only"""
    role = get_user_role(request.user)
    
    if role == 'customer':
        # Customers see only their orders
        orders = Order.objects.filter(customer_user=request.user).select_related('customer_user', 'customer').prefetch_related('items').order_by('-created_at')
    else:
        # Team members see all orders
        orders = Order.objects.filter(order_type='sale').select_related('customer_user', 'customer').prefetch_related('items').order_by('-created_at')
    
    context = {'orders': orders}
    return render(request, 'orders/my_orders.html', context)

@require_permission('orders', 'export')  # Export orders permission
def export_orders(request):
    format_type = request.GET.get('format', 'excel')
    order_type = request.GET.get('type', 'sale')
    
    orders = Order.objects.filter(order_type=order_type).order_by('-created_at')
    
    # Apply same filters as main view
    status_filter = request.GET.get('status')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if status_filter and status_filter != 'all':
        orders = orders.filter(status=status_filter)
    if from_date:
        orders = orders.filter(created_at__date__gte=from_date)
    if to_date:
        orders = orders.filter(created_at__date__lte=to_date)
    
    if format_type == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f'{order_type.title()} Orders'
        
        # Headers
        headers = ['Order Number', 'Customer', 'Date', 'Items', 'Total Amount', 'Status', 'Payment Status']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='0014A8', end_color='0014A8', fill_type='solid')
        
        # Data
        for order in orders:
            ws.append([
                order.order_number,
                order.customer.customer_name if order.customer else 'N/A',
                order.created_at.strftime('%Y-%m-%d'),
                order.items.count(),
                float(order.total_amount or 0),
                order.get_status_display(),
                order.get_payment_status_display() if hasattr(order, 'get_payment_status_display') else 'N/A'
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={order_type}_orders.xlsx'
        wb.save(response)
        return response
    
    else:  # CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={order_type}_orders.csv'
        
        writer = csv.writer(response)
        writer.writerow(['Order Number', 'Customer', 'Date', 'Items', 'Total Amount', 'Status', 'Payment Status'])
        
        for order in orders:
            writer.writerow([
                order.order_number,
                order.customer.customer_name if order.customer else 'N/A',
                order.created_at.strftime('%Y-%m-%d'),
                order.items.count(),
                float(order.total_amount or 0),
                order.get_status_display(),
                order.get_payment_status_display() if hasattr(order, 'get_payment_status_display') else 'N/A'
            ])
        
        return response

@login_required
def sales_orders_api(request):
    orders = Order.objects.filter(order_type='sale').select_related('customer_user', 'customer').prefetch_related('items').order_by('-created_at')
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    if search:
        orders = orders.filter(order_number__icontains=search)
    if status:
        orders = orders.filter(status=status)
    if from_date:
        orders = orders.filter(created_at__date__gte=from_date)
    if to_date:
        orders = orders.filter(created_at__date__lte=to_date)
    
    data = [{
        'id': o.id,
        'order_number': o.order_number,
        'customer_name': o.customer_user.username if o.customer_user else (o.customer.customer_name if o.customer else 'N/A'),
        'date': o.created_at.strftime('%Y-%m-%d'),
        'items_count': o.items.count(),
        'total': float(o.total_amount or 0),
        'status': o.get_status_display()
    } for o in orders[:100]]
    
    return JsonResponse({'orders': data})

@login_required
def purchase_orders_api(request):
    orders = Order.objects.filter(order_type='purchase').select_related('supplier').prefetch_related('items').order_by('-created_at')
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    
    if search:
        orders = orders.filter(order_number__icontains=search)
    if status:
        orders = orders.filter(status=status)
    if from_date:
        orders = orders.filter(created_at__date__gte=from_date)
    if to_date:
        orders = orders.filter(created_at__date__lte=to_date)
    
    data = [{
        'id': o.id,
        'po_number': o.order_number,
        'supplier_name': o.supplier.supplier_name if o.supplier else 'N/A',
        'date': o.created_at.strftime('%Y-%m-%d'),
        'items_count': o.items.count(),
        'total': float(o.total_amount or 0),
        'status': o.get_status_display()
    } for o in orders[:100]]
    
    return JsonResponse({'orders': data})

@require_permission('orders', 'export')  # Download single order
def download_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    # Create Excel file for single order
    wb = Workbook()
    ws = wb.active
    ws.title = f'Order {order.order_number}'
    
    # Order header info
    ws.append(['Order Details'])
    ws.append(['Order Number:', order.order_number])
    ws.append(['Customer:', order.customer.customer_name if order.customer else 'N/A'])
    ws.append(['Date:', order.created_at.strftime('%Y-%m-%d %H:%M')])
    ws.append(['Status:', order.get_status_display()])
    ws.append(['Total Amount:', f'₹{order.total_amount}'])
    ws.append([])
    
    # Order items header
    ws.append(['Order Items'])
    ws.append(['Product', 'Quantity', 'Unit Price', 'Total Price'])
    
    # Order items data
    for item in order.items.all():
        ws.append([
            item.product_name,
            item.quantity,
            f'₹{item.unit_price}',
            f'₹{item.total_price}'
        ])
    
    # Style headers
    for row in [1, 8, 9]:
        for cell in ws[row]:
            cell.font = Font(bold=True)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=order_{order.order_number}.xlsx'
    wb.save(response)
    return response