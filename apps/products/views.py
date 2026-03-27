from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.views.decorators.http import require_http_methods
from django.core.files.storage import default_storage
from goods.models import ListModel as Product
from supplier.models import ListModel as Supplier
from stock.models import StockListModel, StockMovement
from permissions.decorators import require_permission, require_role
import json
import csv
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

@require_role('superadmin', 'admin', 'subadmin', 'staff')
def products_list(request):
    """Unified Products & Inventory page for team roles only"""
    from permissions.decorators import check_permission, get_user_role
    
    user_role = get_user_role(request.user)
    
    # Get user permissions for template
    user_permissions = {
        'products': {
            'view': check_permission(request.user, 'products', 'view'),
            'create': check_permission(request.user, 'products', 'create'),
            'edit': check_permission(request.user, 'products', 'edit'),
            'delete': check_permission(request.user, 'products', 'delete'),
            'export': check_permission(request.user, 'products', 'export'),
            'import': check_permission(request.user, 'products', 'import'),
        }
    }
    
    # Create simplified permissions object for JavaScript
    can_create = user_permissions['products']['create']
    can_edit = user_permissions['products']['edit']
    can_delete = user_permissions['products']['delete']
    can_export = user_permissions['products']['export']
    can_import = user_permissions['products']['import']
    
    return render(request, 'products/unified_products_inventory.html', {
        'user_permissions': user_permissions,
        'user_role': user_role,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_export': can_export,
        'can_import': can_import,
    })

@login_required
def customer_marketplace(request):
    """Marketplace view - shows different content based on role"""
    from permissions.decorators import get_user_role
    
    user_role = get_user_role(request.user)
    
    # Team roles (superadmin, admin, subadmin, staff) see admin marketplace with products from database
    if user_role in ['superadmin', 'admin', 'subadmin', 'staff']:
        # Get products from database
        products = Product.objects.filter(is_delete=False).order_by('-create_time')[:100]
        
        # Attach stock information and images
        products_with_stock = []
        for product in products:
            stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
            product.stock_qty = stock.goods_qty if stock else 0
            product.stock_available = stock.can_order_stock if stock else 0
            product.stock_image = None
            if stock and stock.goods_image:
                try:
                    if hasattr(stock.goods_image, 'url'):
                        product.stock_image = stock.goods_image.url
                    else:
                        product.stock_image = f'/media/{stock.goods_image}'
                except:
                    pass
            products_with_stock.append(product)
        
        # Get categories and suppliers for filters
        categories = Product.objects.filter(is_delete=False).exclude(goods_class__isnull=True).exclude(goods_class='').values_list('goods_class', flat=True).distinct()
        suppliers = Product.objects.filter(is_delete=False).exclude(goods_supplier__isnull=True).exclude(goods_supplier='').values_list('goods_supplier', flat=True).distinct()
        
        return render(request, 'products/admin_marketplace.html', {
            'user_role': user_role,
            'products': products_with_stock,
            'total_products': len(products_with_stock),
            'categories': categories,
            'suppliers': suppliers,
        })
    
    # Customers and guests see customer marketplace
    return render(request, 'products/customer_marketplace.html')

@login_required
def product_detail_page(request, product_id):
    """Product detail page"""
    product = get_object_or_404(Product, id=product_id, is_delete=False)
    stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
    
    return render(request, 'products/detail.html', {
        'product': product,
        'stock': stock
    })

def products_api(request):
    """Products API - public read, auth for write"""
    try:
        # Allow public read access (for guests), but check permissions for write
        if request.method in ['POST', 'PUT', 'DELETE']:
            if not request.user.is_authenticated:
                return JsonResponse({'error': 'Authentication required'}, status=401)
            
            # RBAC checks for write operations
            from permissions.decorators import check_permission, get_user_role
            user_role = get_user_role(request.user)
            
            if request.method == 'POST':
                # Only SuperAdmin and Admin can create products
                if user_role not in ['superadmin', 'admin']:
                    return JsonResponse({'error': 'Permission denied. Only Admin and SuperAdmin can create products.'}, status=403)
                if not check_permission(request.user, 'products', 'create'):
                    return JsonResponse({'error': 'Permission denied'}, status=403)
            elif request.method == 'PUT':
                # SubAdmin can edit, Admin and SuperAdmin can edit
                if user_role not in ['superadmin', 'admin', 'subadmin']:
                    return JsonResponse({'error': 'Permission denied'}, status=403)
                if not check_permission(request.user, 'products', 'edit'):
                    return JsonResponse({'error': 'Permission denied'}, status=403)
            elif request.method == 'DELETE':
                # Only SuperAdmin and Admin can delete
                if user_role not in ['superadmin', 'admin']:
                    return JsonResponse({'error': 'Permission denied. Only Admin and SuperAdmin can delete products.'}, status=403)
                if not check_permission(request.user, 'products', 'delete'):
                    return JsonResponse({'error': 'Permission denied'}, status=403)
        
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('per_page', 24))
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        supplier = request.GET.get('supplier', '')
        stock_filter = request.GET.get('stock', '')
        min_price = request.GET.get('min_price', '')
        max_price = request.GET.get('max_price', '')
        sort_by = request.GET.get('sort', '-create_time')
        
        products = Product.objects.filter(is_delete=False)
        
        # Search filter
        if search:
            products = products.filter(
                Q(goods_code__icontains=search) |
                Q(goods_desc__icontains=search)
            )
        
        # Category filter
        if category:
            products = products.filter(goods_class=category)
        
        # Supplier filter
        if supplier:
            products = products.filter(goods_supplier=supplier)
        
        # Price range filter
        if min_price:
            products = products.filter(goods_price__gte=float(min_price))
        if max_price:
            products = products.filter(goods_price__lte=float(max_price))
        
        # Stock filter
        if stock_filter == 'low':
            stock_codes = StockListModel.objects.filter(goods_qty__lt=10).values_list('goods_code', flat=True)
            products = products.filter(goods_code__in=stock_codes)
        elif stock_filter == 'out':
            stock_codes = StockListModel.objects.filter(goods_qty=0).values_list('goods_code', flat=True)
            products = products.filter(goods_code__in=stock_codes)
        elif stock_filter == 'normal':
            stock_codes = StockListModel.objects.filter(goods_qty__gte=10).values_list('goods_code', flat=True)
            products = products.filter(goods_code__in=stock_codes)
        
        products = products.order_by(sort_by)
        paginator = Paginator(products, page_size)
        page_obj = paginator.get_page(page)
        
        results = []
        for product in page_obj:
            stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
            stock_qty = int(stock.goods_qty) if stock else 0
            image_url = None
            try:
                if stock and stock.goods_image:
                    if hasattr(stock.goods_image, 'url'):
                        image_url = stock.goods_image.url
                    else:
                        # Handle string path
                        image_url = f'/media/{stock.goods_image}'
            except Exception:
                image_url = None
            
            # Get category created_by info
            category_created_by = None
            category_created_at = None
            try:
                from categories.models import Category
                category_obj = Category.objects.filter(name=product.goods_class).first()
                if category_obj and category_obj.created_by:
                    category_created_by = category_obj.created_by.username
                    category_created_at = category_obj.created_at.strftime('%b %d, %Y')
            except:
                pass
            
            results.append({
                'id': product.id,
                'goods_code': product.goods_code,
                'goods_name': product.goods_desc,
                'goods_desc': product.goods_desc,
                'goods_price': float(product.goods_price) if product.goods_price else 0,
                'goods_class': product.goods_class or 'General',
                'goods_supplier': product.goods_supplier or 'N/A',
                'stock': stock_qty,
                'stock_status': 'danger' if stock_qty == 0 else 'warning' if stock_qty < 10 else 'success',
                'status': 'Active' if not product.is_delete else 'Inactive',
                'image': image_url or '/static/images/no-image.svg',
                'is_multistock': True if getattr(product, 'goods_brand', '') == 'Multistock' else False,
                'category_created_by': category_created_by,
                'category_created_at': category_created_at
            })
        
        return JsonResponse({
            'results': results,
            'count': paginator.count,
            'next': page < paginator.num_pages,
            'previous': page > 1,
            'current_page': page,
            'total_pages': paginator.num_pages
        })
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Products API Error: {error_detail}")  # Log to console
        return JsonResponse({
            'error': str(e),
            'detail': error_detail if request.user.is_superuser else 'An error occurred'
        }, status=400)

@login_required
def product_detail_page(request, product_id):
    """Product detail page"""
    product = get_object_or_404(Product, id=product_id, is_delete=False)
    stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
    
    return render(request, 'products/detail.html', {
        'product': product,
        'stock': stock
    })

@require_permission('products', 'edit')  # Edit product fields
def update_product_field(request):
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_id = data.get('id')
            field = data.get('field')
            value = data.get('value')
            
            product = Product.objects.get(id=product_id)
            if field == 'price':
                product.goods_price = value
            elif field == 'stock':
                stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
                if stock:
                    stock.goods_qty = value
                    stock.save()
            product.save()
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'success': False}, status=400)
    return JsonResponse({'error': 'POST required'}, status=405)

@login_required
@require_permission('products', 'edit')
def bulk_update_products(request):
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_ids = data.get('ids', [])
            action = data.get('action')
            value = data.get('value')
            
            products = Product.objects.filter(id__in=product_ids)
            
            if action == 'update_price':
                for product in products:
                    adjustment = float(value)
                    if adjustment > 0:
                        product.goods_price = float(product.goods_price or 0) * (1 + adjustment/100)
                    else:
                        product.goods_price = float(product.goods_price or 0) * (1 + adjustment/100)
                    product.save()
            elif action == 'update_category':
                products.update(goods_class=value)
            elif action == 'update_stock':
                adjustment = int(value)
                for product in products:
                    stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
                    if stock:
                        stock.goods_qty = max(0, stock.goods_qty + adjustment)
                        stock.save()
            
            return JsonResponse({'success': True, 'updated': len(product_ids)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'POST required'}, status=405)

@login_required
@require_permission('products', 'delete')
def bulk_delete_products(request):
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_ids = data.get('ids', [])
            Product.objects.filter(id__in=product_ids).update(is_delete=True)
            return JsonResponse({'success': True, 'deleted': len(product_ids)})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'error': 'POST required'}, status=405)

@require_permission('products', 'view')  # Analytics view permission
def product_analytics_api(request):
    try:
        products = Product.objects.filter(is_delete=False)
        total_products = products.count()
        total_value = 0
        low_stock = 0
        out_of_stock = 0
        
        for product in products:
            stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
            if stock:
                total_value += stock.goods_qty * float(product.goods_price or 0)
                if stock.goods_qty == 0:
                    out_of_stock += 1
                elif stock.goods_qty < 10:
                    low_stock += 1
        
        suppliers = Supplier.objects.filter(is_delete=False).count()
        categories = products.values('goods_class').distinct().count()
        
        return JsonResponse({
            'total_products': total_products,
            'total_value': round(total_value, 2),
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            'active_suppliers': suppliers,
            'categories': categories
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
@require_permission('products', 'create')
@require_http_methods(["POST"])
def create_product(request):
    """Create product - Only SuperAdmin and Admin can create"""
    from permissions.decorators import get_user_role
    user_role = get_user_role(request.user)
    if user_role not in ['superadmin', 'admin']:
        return JsonResponse({'success': False, 'error': 'Permission denied. Only Admin and SuperAdmin can create products.'}, status=403)
    try:
        data = json.loads(request.body)
        
        product = Product.objects.create(
            goods_code=data['sku'],
            goods_desc=data.get('name', data.get('description', '')),
            goods_price=data['price'],
            goods_class=data.get('category', 'General'),
            goods_supplier=data.get('supplier', ''),
            goods_unit='pcs',
            goods_brand='Multistock' if data.get('is_multistock') == 'true' else '',
            goods_weight=0,
            goods_w=0,
            goods_d=0,
            goods_h=0
        )
        
        # Create initial stock
        if 'stock' in data:
            StockListModel.objects.create(
                goods_code=product.goods_code,
                goods_desc=product.goods_desc,
                goods_qty=data['stock'],
                onhand_stock=data['stock'],
                can_order_stock=data['stock']
            )
        
        return JsonResponse({
            'success': True,
            'id': product.id,
            'message': 'Product created successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('products', 'view')  # View product details
def product_detail_api(request, product_id):
    try:
        product = get_object_or_404(Product, id=product_id, is_delete=False)
        stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
        
        # Get image URL
        image_url = None
        if stock and stock.goods_image:
            try:
                if hasattr(stock.goods_image, 'url'):
                    image_url = stock.goods_image.url
                else:
                    image_url = f'/media/{stock.goods_image}'
            except:
                image_url = str(stock.goods_image)
        
        return JsonResponse({
            'id': product.id,
            'goods_code': product.goods_code,
            'goods_name': product.goods_desc,
            'goods_desc': product.goods_desc,
            'goods_price': float(product.goods_price) if product.goods_price else 0,
            'goods_class': product.goods_class or 'General',
            'goods_supplier': product.goods_supplier or 'N/A',
            'goods_unit': product.goods_unit or 'pcs',
            'goods_brand': product.goods_brand or 'N/A',
            'goods_weight': float(product.goods_weight) if product.goods_weight else 0,
            'goods_image': image_url,
            'stock': int(stock.goods_qty) if stock else 0,
            'created_at': product.create_time.strftime('%Y-%m-%d %H:%M:%S') if product.create_time else '',
            'updated_at': product.update_time.strftime('%Y-%m-%d %H:%M:%S') if product.update_time else ''
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@require_permission('products', 'edit')  # Update products
@require_http_methods(["PUT", "POST"])
def update_product(request, product_id):
    """Update product - SuperAdmin, Admin, and Supervisor can edit"""
    from permissions.decorators import get_user_role
    user_role = get_user_role(request.user)
    if user_role not in ['superadmin', 'admin', 'supervisor']:
        return JsonResponse({'success': False, 'error': 'Permission denied. Only Admin, SuperAdmin, and Supervisor can edit products.'}, status=403)
    try:
        product = get_object_or_404(Product, id=product_id)
        
        # Handle form data (for file uploads)
        if request.content_type.startswith('multipart/form-data'):
            data = request.POST.dict()
            image_file = request.FILES.get('image')
            image_url = data.get('image_url', '').strip()
        else:
            data = json.loads(request.body)
            image_file = None
            image_url = data.get('image_url', '').strip()
        
        if 'name' in data:
            product.goods_desc = data['name']
        if 'description' in data:
            product.goods_desc = data['description']
        if 'price' in data:
            product.goods_price = data['price']
        if 'category' in data:
            product.goods_class = data['category']
        if 'supplier' in data:
            product.goods_supplier = data['supplier']
        if 'is_multistock' in data:
            product.goods_brand = 'Multistock' if data['is_multistock'] == 'true' else product.goods_brand
        
        product.save()
        
        # Update stock and image
        stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
        if not stock:
            stock = StockListModel.objects.create(
                goods_code=product.goods_code,
                goods_desc=product.goods_desc,
                goods_qty=0,
                onhand_stock=0,
                can_order_stock=0
            )
        
        if 'stock' in data:
            stock.goods_qty = int(data['stock'])
            stock.onhand_stock = int(data['stock'])
            stock.can_order_stock = int(data['stock'])
        
        # Handle image update
        import os
        from django.conf import settings
        import urllib.request
        import time
        
        if image_file or image_url:
            # Delete old image if exists
            if stock.goods_image:
                try:
                    old_path = os.path.join(settings.MEDIA_ROOT, str(stock.goods_image))
                    if os.path.exists(old_path):
                        os.remove(old_path)
                except:
                    pass
            
            if image_file:
                stock.goods_image = image_file
            elif image_url:
                try:
                    # Download image from URL
                    img_dir = os.path.join(settings.MEDIA_ROOT, 'products')
                    os.makedirs(img_dir, exist_ok=True)
                    img_path = os.path.join(img_dir, f"{product.goods_code}_{int(time.time())}.jpg")
                    urllib.request.urlretrieve(image_url, img_path)
                    stock.goods_image = f"products/{product.goods_code}_{int(time.time())}.jpg"
                except Exception as e:
                    return JsonResponse({'success': False, 'error': f'Failed to download image: {str(e)}'}, status=400)
        
        stock.save()
        
        return JsonResponse({'success': True, 'message': 'Product updated'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('products', 'export')  # Export products
def export_products(request):
    try:
        format_type = request.GET.get('format', 'excel')
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        
        products = Product.objects.filter(is_delete=False)
        
        if search:
            products = products.filter(
                Q(goods_code__icontains=search) |
                Q(goods_desc__icontains=search)
            )
        if category:
            products = products.filter(goods_class=category)
        
        if format_type == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products'
            
            # Headers
            headers = ['SKU', 'Name', 'Description', 'Price', 'Category', 'Supplier', 'Stock']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='0014A8', end_color='0014A8', fill_type='solid')
            
            # Data
            for product in products:
                stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
                ws.append([
                    product.goods_code,
                    product.goods_desc,
                    product.goods_desc,
                    float(product.goods_price) if product.goods_price else 0,
                    product.goods_class,
                    product.goods_supplier,
                    int(stock.goods_qty) if stock else 0
                ])
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=products.xlsx'
            wb.save(response)
            return response
        else:
            # CSV export
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=products.csv'
            
            writer = csv.writer(response)
            writer.writerow(['SKU', 'Name', 'Description', 'Price', 'Category', 'Supplier', 'Stock'])
            
            for product in products:
                stock = StockListModel.objects.filter(goods_code=product.goods_code).first()
                writer.writerow([
                    product.goods_code,
                    product.goods_desc,
                    product.goods_desc,
                    float(product.goods_price) if product.goods_price else 0,
                    product.goods_class,
                    product.goods_supplier,
                    int(stock.goods_qty) if stock else 0
                ])
            
            return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@require_permission('products', 'view')  # Get categories
def get_categories_api(request):
    """Get all unique categories with product counts"""
    categories = Product.objects.filter(is_delete=False).values('goods_class').annotate(
        product_count=Count('id')
    ).order_by('goods_class')
    
    category_list = [{
        'name': cat['goods_class'] or 'Uncategorized',
        'count': cat['product_count']
    } for cat in categories if cat['goods_class']]
    
    return JsonResponse({'categories': category_list})

@require_permission('products', 'view')  # Get suppliers
def get_suppliers_api(request):
    suppliers = Supplier.objects.filter(is_delete=False).values('id', 'supplier_name')
    return JsonResponse({'suppliers': list(suppliers)})

@require_permission('categories', 'create')  # Create category
@require_http_methods(["POST"])
def create_category(request):
    """Create a new category"""
    try:
        data = json.loads(request.body)
        category_name = data.get('name', '').strip()
        
        if not category_name:
            return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)
        
        # Check if category already exists
        exists = Product.objects.filter(is_delete=False, goods_class__iexact=category_name).exists()
        if exists:
            return JsonResponse({'success': False, 'error': 'Category already exists'}, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Category "{category_name}" created successfully',
            'category': category_name
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('categories', 'edit')  # Rename category
@require_http_methods(["POST"])
def rename_category(request):
    """Rename an existing category"""
    try:
        data = json.loads(request.body)
        old_name = data.get('old_name', '').strip()
        new_name = data.get('new_name', '').strip()
        
        if not old_name or not new_name:
            return JsonResponse({'success': False, 'error': 'Both old and new names are required'}, status=400)
        
        # Check if new category name already exists
        exists = Product.objects.filter(is_delete=False, goods_class__iexact=new_name).exclude(goods_class__iexact=old_name).exists()
        if exists:
            return JsonResponse({'success': False, 'error': 'New category name already exists'}, status=400)
        
        # Update all products with the old category
        updated = Product.objects.filter(is_delete=False, goods_class=old_name).update(goods_class=new_name)
        
        return JsonResponse({
            'success': True,
            'message': f'Category renamed from "{old_name}" to "{new_name}"',
            'updated_products': updated
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('categories', 'delete')  # Delete category
@require_http_methods(["POST"])
def delete_category(request):
    """Delete a category (sets products to Uncategorized)"""
    try:
        data = json.loads(request.body)
        category_name = data.get('name', '').strip()
        
        if not category_name:
            return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)
        
        # Update products to 'Uncategorized' instead of deleting
        updated = Product.objects.filter(is_delete=False, goods_class=category_name).update(goods_class='Uncategorized')
        
        return JsonResponse({
            'success': True,
            'message': f'Category "{category_name}" deleted. {updated} products moved to Uncategorized',
            'updated_products': updated
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('categories', 'edit')  # Merge categories
@require_http_methods(["POST"])
def merge_categories(request):
    """Merge multiple categories into one"""
    try:
        data = json.loads(request.body)
        source_categories = data.get('source', [])  # List of categories to merge
        target_category = data.get('target', '').strip()  # Category to merge into
        
        if not source_categories or not target_category:
            return JsonResponse({'success': False, 'error': 'Source and target categories are required'}, status=400)
        
        # Update all products from source categories to target
        updated = Product.objects.filter(is_delete=False, goods_class__in=source_categories).update(goods_class=target_category)
        
        return JsonResponse({
            'success': True,
            'message': f'{len(source_categories)} categories merged into "{target_category}"',
            'updated_products': updated
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
@require_permission('products', 'create')
@require_http_methods(["POST"])
@login_required
@require_permission('products', 'import')
def bulk_import_products(request):
    """Import products from CSV/Excel file"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No file uploaded'}, status=400)
        
        file = request.FILES['file']
        file_ext = file.name.split('.')[-1].lower()
        
        imported = 0
        errors = []
        
        if file_ext == 'csv':
            # Handle CSV
            decoded_file = file.read().decode('utf-8')
            csv_data = csv.DictReader(io.StringIO(decoded_file))
            
            for row in csv_data:
                try:
                    # Create or update product
                    product, created = Product.objects.update_or_create(
                        goods_code=row.get('sku', row.get('SKU', '')),
                        defaults={
                            'goods_desc': row.get('name', row.get('product_name', '')),
                            'goods_price': float(row.get('price', 0)),
                            'goods_class': row.get('category', 'General'),
                            'goods_supplier': row.get('supplier', ''),
                            'goods_unit': row.get('unit', 'pcs'),
                            'is_delete': False
                        }
                    )
                    
                    # Create or update stock
                    stock_qty = int(row.get('stock', row.get('quantity', 0)))
                    StockListModel.objects.update_or_create(
                        goods_code=product.goods_code,
                        defaults={
                            'goods_desc': product.goods_desc,
                            'goods_qty': stock_qty,
                            'can_order_stock': stock_qty,
                            'onhand_stock': stock_qty,
                            'supplier': product.goods_supplier
                        }
                    )
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {imported + 1}: {str(e)}")
        
        elif file_ext in ['xlsx', 'xls']:
            # Handle Excel
            wb = load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = dict(zip(headers, row))
                    
                    # Create or update product
                    product, created = Product.objects.update_or_create(
                        goods_code=row_data.get('sku', row_data.get('SKU', '')),
                        defaults={
                            'goods_desc': row_data.get('name', row_data.get('product_name', '')),
                            'goods_price': float(row_data.get('price', 0) or 0),
                            'goods_class': row_data.get('category', 'General'),
                            'goods_supplier': row_data.get('supplier', ''),
                            'goods_unit': row_data.get('unit', 'pcs'),
                            'is_delete': False
                        }
                    )
                    
                    # Create or update stock
                    stock_qty = int(row_data.get('stock', row_data.get('quantity', 0)) or 0)
                    StockListModel.objects.update_or_create(
                        goods_code=product.goods_code,
                        defaults={
                            'goods_desc': product.goods_desc,
                            'goods_qty': stock_qty,
                            'can_order_stock': stock_qty,
                            'onhand_stock': stock_qty,
                            'supplier': product.goods_supplier
                        }
                    )
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
        else:
            return JsonResponse({'success': False, 'error': 'Invalid file format. Use CSV or Excel'}, status=400)
        
        return JsonResponse({
            'success': True,
            'imported': imported,
            'errors': errors[:10]  # Return first 10 errors
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('products', 'delete')  # Delete product
@require_http_methods(["POST"])
def delete_product(request, product_id):
    """Delete product - Only SuperAdmin and Admin can delete"""
    from permissions.decorators import get_user_role
    user_role = get_user_role(request.user)
    if user_role not in ['superadmin', 'admin']:
        return JsonResponse({'success': False, 'error': 'Permission denied. Only Admin and SuperAdmin can delete products.'}, status=403)
    
    try:
        product = get_object_or_404(Product, id=product_id)
        product.is_delete = True
        product.save()
        return JsonResponse({'success': True, 'message': 'Product deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@require_permission('products', 'view')  # Download import template
def download_import_template(request):
    """Download CSV/Excel template for bulk import"""
    format_type = request.GET.get('format', 'csv')
    
    if format_type == 'excel':
        # Create Excel template
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Import Template"
        
        headers = ['sku', 'name', 'price', 'category', 'supplier', 'stock', 'unit']
        ws.append(headers)
        
        # Add example row
        ws.append(['PROD-001', 'Sample Product', '99.99', 'Electronics', 'Sample Supplier', '100', 'pcs'])
        
        # Style headers
        header_fill = PatternFill(start_color="0014A8", end_color="0014A8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=product_import_template.xlsx'
        wb.save(response)
        return response
    
    else:
        # Create CSV template
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=product_import_template.csv'
        
        writer = csv.writer(response)
        writer.writerow(['sku', 'name', 'price', 'category', 'supplier', 'stock', 'unit'])
        writer.writerow(['PROD-001', 'Sample Product', '99.99', 'Electronics', 'Sample Supplier', '100', 'pcs'])
        
        return response

