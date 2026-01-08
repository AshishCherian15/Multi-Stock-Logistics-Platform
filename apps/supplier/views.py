from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated, BasePermission
from .models import ListModel
from .serializers import SupplierSerializer
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db import models
from orders.models import Order
import csv
import json

# Import RBAC decorators
from permissions.decorators import require_permission, check_permission

# Custom DRF permission class for supplier CRUD
class SupplierPermission(BasePermission):
    """
    RBAC permission class for Supplier ViewSet
    - GET (list/retrieve): suppliers.view permission
    - POST (create): suppliers.create permission
    - PUT/PATCH (update): suppliers.edit permission
    - DELETE: suppliers.delete permission
    """
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        
        if request.method == 'GET':
            return check_permission(request.user, 'suppliers', 'view')
        elif request.method == 'POST':
            return check_permission(request.user, 'suppliers', 'create')
        elif request.method in ['PUT', 'PATCH']:
            return check_permission(request.user, 'suppliers', 'edit')
        elif request.method == 'DELETE':
            return check_permission(request.user, 'suppliers', 'delete')
        return False

class SupplierViewSet(viewsets.ModelViewSet):
    queryset = ListModel.objects.filter(is_delete=False)
    serializer_class = SupplierSerializer
    permission_classes = [SupplierPermission]  # RBAC permission
    
    def get_queryset(self):
        """SuperAdmin sees all suppliers, others see only their store's suppliers."""
        user = self.request.user
        if user.is_superuser:
            return ListModel.objects.filter(is_delete=False)
        
        try:
            from permissions.decorators import get_user_role
            role = get_user_role(user)
            if role == 'superadmin':
                return ListModel.objects.filter(is_delete=False)
        except:
            pass
        
        openid = getattr(user, 'openid', user.username)
        return ListModel.objects.filter(openid=openid, is_delete=False)
    
    def perform_create(self, serializer):
        """Set openid on supplier creation"""
        openid = getattr(self.request.user, 'openid', self.request.user.username)
        serializer.save(openid=openid)
    
    def perform_destroy(self, instance):
        """Admin can only soft delete suppliers."""
        user = self.request.user
        if user.is_superuser:
            instance.delete()  # Hard delete
        else:
            try:
                from permissions.decorators import get_user_role
                if get_user_role(user) == 'superadmin':
                    instance.delete()
                else:
                    instance.is_delete = True  # Soft delete
                    instance.save()
            except:
                instance.is_delete = True
                instance.save()


@require_permission('suppliers', 'view')  # Supplier view permission
def supplier_management_view(request):
    # Filter suppliers by store for non-SuperAdmin
    from permissions.decorators import get_user_role
    if request.user.is_superuser or get_user_role(request.user) == 'superadmin':
        suppliers = ListModel.objects.filter(is_delete=False).order_by('supplier_name')
    else:
        openid = getattr(request.user, 'openid', request.user.username)
        suppliers = ListModel.objects.filter(openid=openid, is_delete=False).order_by('supplier_name')
    
    total_suppliers = suppliers.count()
    active_suppliers = total_suppliers  # all non-deleted suppliers are active
    pending_orders = Order.objects.filter(order_type='purchase', status='pending', supplier__isnull=False).count()
    avg_level = suppliers.aggregate(models.Avg('supplier_level'))['supplier_level__avg'] or 0

    context = {
        'suppliers': suppliers,
        'total_suppliers': total_suppliers,
        'active_suppliers': active_suppliers,
        'pending_orders': pending_orders,
        'avg_level': avg_level,
        'access_denied': False,
        'readonly': not check_permission(request.user, 'suppliers', 'create'),
    }
    return render(request, 'suppliers/enhanced_suppliers.html', context)


@require_permission('suppliers', 'view')
def export_suppliers(request):
    from datetime import datetime
    from permissions.decorators import get_user_role
    
    format_type = request.GET.get('format', 'csv')
    
    if request.user.is_superuser or get_user_role(request.user) == 'superadmin':
        suppliers = ListModel.objects.filter(is_delete=False)
    else:
        openid = getattr(request.user, 'openid', request.user.username)
        suppliers = ListModel.objects.filter(openid=openid, is_delete=False)
    
    if format_type == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Suppliers'
            
            header_fill = PatternFill(start_color='0014A8', end_color='0014A8', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            headers = ['S.No', 'Supplier Name', 'Manager', 'Contact', 'City', 'Address', 'Level', 'Created Date']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            for idx, s in enumerate(suppliers, 1):
                ws.append([
                    idx,
                    s.supplier_name,
                    s.supplier_manager or 'N/A',
                    s.supplier_contact or 'N/A',
                    s.supplier_city or 'N/A',
                    s.supplier_address or 'N/A',
                    s.supplier_level,
                    s.create_time.strftime('%Y-%m-%d %H:%M')
                ])
            
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="suppliers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            format_type = 'csv'
    
    # CSV Export
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="suppliers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['S.No', 'Supplier Name', 'Manager', 'Contact', 'City', 'Address', 'Level', 'Created Date'])
    
    for idx, s in enumerate(suppliers, 1):
        writer.writerow([
            idx,
            s.supplier_name,
            s.supplier_manager or 'N/A',
            s.supplier_contact or 'N/A',
            s.supplier_city or 'N/A',
            s.supplier_address or 'N/A',
            s.supplier_level,
            s.create_time.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response